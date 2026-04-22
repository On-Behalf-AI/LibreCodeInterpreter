#!/usr/bin/env python3
"""
Automated agent tests via LibreChat Open Responses API.

Usage:
    python3 tests/agent_api_tests.py                    # Run all tests
    python3 tests/agent_api_tests.py D01b D01c P01b     # Run specific tests
    python3 tests/agent_api_tests.py --agent docx        # Run all DOCX tests
    python3 tests/agent_api_tests.py --list              # List all tests
    python3 tests/agent_api_tests.py --no-file           # Skip tests that need files
    python3 tests/agent_api_tests.py --generate-files    # Generate sample files

Prerequisites:
    - LibreChat running on localhost:3080
    - Agent API key in AGENT_API_KEY env var or .agent-api-key file
    - All agents deployed with current instructions
    - Test input files in docs/Test_files/ (for NEEDS_FILE tests)

Each test:
    1. Optionally uploads input files via LibreChat API
    2. Sends a prompt to the agent via /api/agents/v1/responses
    3. Checks that the response completed without error
    4. Validates methodology (checks for expected scripts/patterns in reasoning/message)
    5. Produces a PASS/FAIL report
"""

import argparse
import json
import os
import re
import sys
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

try:
    import requests
except ImportError:
    print("pip install requests")
    sys.exit(1)


# === Configuration ===

API_BASE = os.environ.get("LIBRECHAT_URL", "http://127.0.0.1:3080")
API_KEY = os.environ.get("AGENT_API_KEY", "")
CI_BASE = os.environ.get("CODE_INTERPRETER_URL", "http://127.0.0.1:8010")
CI_KEY = os.environ.get("CODE_INTERPRETER_KEY", "REDACTED_KEY_ROTATED_2026-04-23")
JWT_TOKEN = os.environ.get("LIBRECHAT_JWT", "")  # For file uploads (optional)
RESPONSES_ENDPOINT = f"{API_BASE}/api/agents/v1/responses"
RESULTS_DIR = Path("tests/results")
FILES_DIR = Path("tests/results/files")
TEST_FILES_DIR = Path("docs/Test_files")
TIMEOUT = 600  # 10 minutes max per test

# Force unbuffered output
import functools
print = functools.partial(print, flush=True)


# === Test Input File Mapping ===
# Maps test_id → list of file paths relative to TEST_FILES_DIR.
# Tests listed here require input files. If files are missing, the test
# still runs (methodology-only mode) but logs a warning.

TEST_INPUT_FILES = {
    # DOCX
    "D01": ["1. Word/D01_anonymized.docx"],
    "D03": ["1. Word/D02.docx"],
    "D04": ["1. Word/D04_test_03_proposition_commerciale_variantes.docx"],
    "D06": ["1. Word/D06_a_anonymized.docx", "1. Word/D06_b_anonymized.docx"],
    "D07": ["1. Word/D07.docx"],
    "D09": ["1. Word/D09.doc"],
    "D11": ["1. Word/D11.docx"],
    "D12": ["1. Word/D12.docx"],
    # PPTX
    "P02": ["2. PPT/P02.pptx"],
    "P03": ["2. PPT/P03.pptx"],
    "P04": ["2. PPT/P04.pptx"],
    "P05": ["2. PPT/P05.pptx"],
    "P06": ["2. PPT/P02.pptx"],
    "P07": ["2. PPT/P05.pptx"],
    "P09": ["2. PPT/P09.pptx"],
    "P10": ["2. PPT/P10.pptx"],
    "P12": ["2. PPT/P12.potx"],
    # XLSX
    "X02": ["3. Excel/X02_donnees_commerciales_complexes.xlsx"],
    "X04": ["3. Excel/X04_ventes_mensuelles_input_complexe.xlsx"],
    "X05": ["3. Excel/X05_ancien_format_complexe.xls"],
    "X06": ["3. Excel/X06_transactions_pivot_input_complexe.xlsx"],
    "X08": ["3. Excel/X08_source_export_pdf_complexe.xlsx"],
    "X09": ["3. Excel/X09_rapport_complexe_janvier.xlsx",
             "3. Excel/X09_rapport_complexe_fevrier.xlsx",
             "3. Excel/X09_rapport_complexe_mars.xlsx"],
    "X10": ["3. Excel/X10_donnees_sales_complexes.xlsx"],
    "X12": ["3. Excel/X12_donnees_rh_complexes.xlsx"],
    # PDF
    "F01": ["4. PDF/F01_contrat/service-agreement.pdf"],
    "F02": ["4. PDF/F02_tableaux/invoice-sample.pdf"],
    "F03": ["4. PDF/F03_ocr/scan-sample.pdf"],
    "F05": ["4. PDF/F05_split/document-15pages.pdf"],
    "F06": ["4. PDF/F06_integrite/corrupted.pdf",
             "4. PDF/F06_integrite/encrypted.pdf",
             "4. PDF/F06_integrite/not_encrypted.pdf"],
    "F07": ["4. PDF/F07_images/document-a-convertir.pdf"],
    "F08": ["4. PDF/F08_metadata/invoice-metadata.pdf"],
    "F09": ["4. PDF/F09_rotation/document-pages-retournees.pdf"],
    "F10": ["4. PDF/F10_watermark/document-a-watermarker.pdf"],
    "F11": ["4. PDF/F11_compression/sample-heavy-25mb.pdf"],
    "F12": ["4. PDF/F12_pipeline_ocr/bank-statement-scanned.pdf"],
    # FFmpeg / Quick Edits
    "M01": ["5. Medias/file_example_MOV_640_800kB.mov"],
    "M02": ["5. Medias/sample-10s.mp4"],
    "M03": ["5. Medias/sample_1280x720_surfing_with_audio.avi"],
    "M04": ["5. Medias/sample-10s.mp4"],
    "M07": ["5. Medias/Landscape_big_river_in_mountains.jpg"],
    "M08": ["5. Medias/sample-15s.mp4", "5. Medias/sample-20s.mp4"],
    "M09": ["5. Medias/sample-10s.mp4", "5. Medias/sample-12s.mp3"],
    "M10": ["5. Medias/sample-boat-400x300.png",
             "5. Medias/sample-city-park-400x300.jpg",
             "5. Medias/sample-clouds-400x300.jpg",
             "5. Medias/sample-birch-400x300.jpg"],
    "M11": ["5. Medias/sample_1280x720_surfing_with_audio.avi"],
    "M12": ["5. Medias/sample-boat-400x300.png",
             "5. Medias/sample-city-park-400x300.jpg",
             "5. Medias/sample-clouds-400x300.jpg",
             "5. Medias/sample-clouds2-400x300.png",
             "5. Medias/sample-bumblebee-400x300.png"],
}


@dataclass
class TestResult:
    test_id: str
    agent: str
    status: str = "PENDING"  # PASS, FAIL, ERROR, SKIP
    duration: float = 0
    prompt: str = ""
    response_text: str = ""
    code_blocks: list = field(default_factory=list)
    files_generated: list = field(default_factory=list)
    files_uploaded: list = field(default_factory=list)
    methodology_checks: dict = field(default_factory=dict)
    error: str = ""


@dataclass
class TestCase:
    test_id: str
    agent_id: str
    agent_name: str
    prompt: str
    methodology_patterns: list  # Regex patterns expected in code execution
    methodology_antipatterns: list = field(default_factory=list)
    description: str = ""
    expect_file: bool = False
    file_extension: str = ""
    needs_file: bool = False  # True if test requires input files
    previous_response_id: Optional[str] = None


# === API Client ===

def exec_code(code: str, lang: str = "py") -> dict:
    """Execute code directly in the code-interpreter sandbox."""
    headers = {"x-api-key": CI_KEY, "Content-Type": "application/json"}
    resp = requests.post(f"{CI_BASE}/exec", headers=headers,
                         json={"code": code, "lang": lang}, timeout=TIMEOUT)
    resp.raise_for_status()
    return resp.json()


def download_file(session_id: str, file_id: str, output_path: Path) -> bool:
    """Download a file from the code-interpreter."""
    headers = {"x-api-key": CI_KEY}
    resp = requests.get(f"{CI_BASE}/download/{session_id}/{file_id}",
                        headers=headers, timeout=60)
    if resp.status_code == 200 and len(resp.content) > 0:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_bytes(resp.content)
        return True
    return False


def upload_file(file_path: Path) -> Optional[str]:
    """Upload a file to LibreChat and return file_id.

    Requires JWT_TOKEN (set via --jwt-token or LIBRECHAT_JWT env var).
    LibreChat's /api/files endpoint uses JWT auth, not API key auth.
    The Open Responses API does NOT yet pass uploaded files to agents
    (requestFiles: [] is hardcoded in responses.js). This is future-proofing.
    """
    if not JWT_TOKEN:
        return None

    headers = {"Authorization": f"Bearer {JWT_TOKEN}"}
    with open(file_path, "rb") as f:
        resp = requests.post(
            f"{API_BASE}/api/files",
            headers=headers,
            files={"file": (file_path.name, f, "application/octet-stream")},
            data={"endpoint": "agents"},
            timeout=120,
        )
    if resp.status_code == 200:
        data = resp.json()
        return data.get("file_id") or data.get("id") or data.get("_id")
    return None


def call_agent(agent_id: str, prompt: str, file_ids: list = None,
               previous_response_id: str = None) -> dict:
    """Call an agent via the Open Responses API."""
    headers = {
        "Authorization": f"Bearer {API_KEY}",
        "Content-Type": "application/json",
    }

    # Build input: structured if files attached, simple string otherwise
    if file_ids:
        content = [{"type": "input_text", "text": prompt}]
        for fid in file_ids:
            content.append({"type": "input_file", "file_id": fid})
        input_data = [{"role": "user", "content": content}]
    else:
        input_data = prompt

    payload = {
        "model": agent_id,
        "input": input_data,
        "stream": False,
    }
    if previous_response_id:
        payload["previous_response_id"] = previous_response_id

    resp = requests.post(RESPONSES_ENDPOINT, headers=headers, json=payload, timeout=TIMEOUT)
    resp.raise_for_status()
    return resp.json()


def extract_response_data(response: dict) -> tuple:
    """Extract text, reasoning, and tool calls from response.

    NOTE: The Open Responses API does NOT return the actual code executed
    (function_call.arguments is always empty). We can only check methodology
    via the reasoning text (agent's thinking) and the message text (agent's output).
    """
    text_parts = []
    code_blocks = []
    files = []

    for output in response.get("output", []):
        if output.get("type") == "message":
            for content in output.get("content", []):
                if content.get("type") == "output_text":
                    text = content.get("text", "")
                    text_parts.append(text)
                    code_blocks.append(("message", text))
        elif output.get("type") == "reasoning":
            for content in output.get("content", []):
                if content.get("type") == "reasoning_text":
                    code_blocks.append(("reasoning", content.get("text", "")))
        elif output.get("type") == "function_call":
            name = output.get("name", "")
            args = output.get("arguments", "")
            code_blocks.append(("function_call", f"TOOL_USED: {name} {args}"))

    response_text = "\n".join(text_parts)
    return response_text, code_blocks, files


def check_methodology(code_blocks: list, patterns: list, antipatterns: list) -> dict:
    """Check that expected methodology patterns are present in code execution."""
    all_code = "\n".join(text for _, text in code_blocks)
    results = {}

    for pattern in patterns:
        found = bool(re.search(pattern, all_code, re.IGNORECASE | re.DOTALL))
        results[f"EXPECTED: {pattern}"] = "FOUND" if found else "MISSING"

    for pattern in antipatterns:
        found = bool(re.search(pattern, all_code, re.IGNORECASE | re.DOTALL))
        results[f"FORBIDDEN: {pattern}"] = "VIOLATION" if found else "OK"

    return results


# === File generation scripts (run directly in code-interpreter, bypass LLM) ===

FILE_GENERATION_SCRIPTS = {
    "D01b": '''
import subprocess, json, os
os.chdir('/mnt/data')
config = {"meeting":{"title":"CR Test","subtitle":"Test","date":"17/04/2026","location":"Test","organizer":"Test"},"participants":[{"name":"A","role":"R","company":"C"}],"sections":[{"title":"Test","level":1,"content":[{"type":"text","text":"Contenu test."}]}]}
with open("/tmp/config.json","w") as f: json.dump(config,f,ensure_ascii=False)
subprocess.run(["python3","/opt/skills/docx/scripts/fill_cr_template.py","/opt/skills/docx/templates/corporate/template-compte-rendu.docx","cr_test.docx","/tmp/config.json"],check=True)
print("OK")
''',
    "D01c": '''
import subprocess, json, os
os.chdir('/mnt/data')
config = {"placeholders":{"[TITRE DU DOCUMENT]":"Guide Test","[Sous-titre du document]":"Test","[Auteur]":"Test","[Date]":"17/04/2026"},"sections":[{"title":"Section 1","level":1,"content":[{"type":"text","text":"Contenu."},{"type":"bullets","items":["A","B"]},{"type":"code","text":"echo hello"},{"type":"table","headers":["Col1","Col2"],"rows":[["a","b"]]}]}]}
with open("/tmp/config.json","w") as f: json.dump(config,f,ensure_ascii=False)
subprocess.run(["python3","/opt/skills/docx/scripts/fill_template.py","/opt/skills/docx/templates/corporate/template-base.docx","guide_test.docx","/tmp/config.json"],check=True)
print("OK")
''',
    "P01b": '''
const pptxgen = require("pptxgenjs");
const pptx = new pptxgen();
pptx.layout = "LAYOUT_16x9";
const s = pptx.addSlide();
s.background = { fill: "1C244B" };
s.addText("Test Slide", {x:0.5,y:2,w:9,h:1.5,fontSize:36,fontFace:"Arial",bold:true,color:"FFFFFF"});
pptx.writeFile({fileName:"/mnt/data/test_pptx.pptx"}).then(()=>console.log("OK"));
''',
    "X01": '''
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
wb = Workbook()
ws = wb.active
ws.title = "Budget"
hdr = PatternFill(start_color="2F5597",end_color="2F5597",fill_type="solid")
hfont = Font(name="Arial",size=11,bold=True,color="FFFFFF")
for col,h in enumerate(["Poste","Montant"],1):
    c = ws.cell(row=1,column=col,value=h)
    c.fill = hdr
    c.font = hfont
ws.cell(row=2,column=1,value="Salaires")
ws.cell(row=2,column=2,value=45000)
wb.save("/mnt/data/budget_test.xlsx")
print("OK")
''',
    "A02": '''
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
fig, axes = plt.subplots(2,2,figsize=(10,8))
colors = ["#2F5597","#5B9AD4","#FB840D","#FCA810"]
axes[0,0].plot([1,2,3,4],[10,20,15,25],color=colors[0])
axes[0,0].set_title("CA Mensuel")
axes[0,1].pie([40,30,20,10],labels=["A","B","C","D"],colors=colors)
axes[0,1].set_title("Répartition")
axes[1,0].barh(["C1","C2","C3"],[100,80,60],color=colors[1])
axes[1,0].set_title("Top Clients")
axes[1,1].scatter(np.random.rand(20),np.random.rand(20),color=colors[2])
axes[1,1].set_title("Scatter")
plt.suptitle("Dashboard corporate",fontweight="bold")
plt.tight_layout()
plt.savefig("/mnt/data/dashboard_test.png",dpi=150,bbox_inches="tight")
plt.close()
print("OK")
''',
}


# === Test Definitions ===

TESTS = []


def test(test_id, agent_id, agent_name, prompt, patterns, antipatterns=None,
         description="", expect_file=False, file_ext=""):
    """Register a test case."""
    TESTS.append(TestCase(
        test_id=test_id,
        agent_id=agent_id,
        agent_name=agent_name,
        prompt=prompt,
        methodology_patterns=patterns,
        methodology_antipatterns=antipatterns or [],
        description=description,
        expect_file=expect_file,
        file_extension=file_ext,
        needs_file=test_id in TEST_INPUT_FILES,
    ))


# ==========================================
# DOCX Agent — Self-contained tests
# ==========================================

test("D01b", "agent_docx_complete", "DOCX",
     "Crée un compte-rendu de la réunion suivante : Réunion du 10 avril 2026, visioconférence Teams, "
     "organisée par Jean Dupont. Participants : Sophie Martin (Directrice RH, Nextera Corp) et "
     "Jean Dupont (Consultant IA, Acme Corp). Sujet : cadrage projet IA RH. "
     "Décisions : lancement POC chatbot RH. Actions : étude faisabilité avant le 24 avril.",
     patterns=[
         r"fill_cr_template|compte.rendu|template.*CR",
         r"execute_code",
     ],
     description="Créer un CR via fill_cr_template.py",
     expect_file=True, file_ext=".docx")

test("D01c", "agent_docx_complete", "DOCX",
     "Crée un guide d'installation pour PostgreSQL sur Ubuntu, avec prérequis, étapes d'installation, "
     "configuration de base, et dépannage.",
     patterns=[
         r"fill_template|template.*base|template.*corporate",
         r"execute_code",
     ],
     description="Créer un guide technique via fill_template.py",
     expect_file=True, file_ext=".docx")

test("D02", "agent_docx_complete", "DOCX",
     "Voici un texte de CGV simplifié. Remplace 'le Client' par 'l'Utilisateur' et '30 jours' par "
     "'15 jours ouvrés' en tracked changes.\n\n"
     "Article 1 : le Client s'engage à respecter les présentes conditions. Le paiement est dû sous 30 jours. "
     "Article 2 : le Client peut résilier sous 30 jours de préavis.",
     patterns=[
         r"tracked.replace|tracked.changes|redline",
         r"execute_code",
     ],
     description="Tracked changes avec {{current_user}} comme auteur")

test("D05", "agent_docx_complete", "DOCX",
     "Convertis ce texte markdown en document Word avec les styles corporate :\n\n"
     "# Politique de télétravail\n## 1. Principes généraux\n"
     "Le télétravail est ouvert à tous les collaborateurs.\n"
     "## 2. Modalités\n- Maximum 3 jours par semaine\n- Accord du manager requis\n"
     "## 3. Obligations\nRespecter les horaires définis.",
     patterns=[r"pandoc|markdown|convert", r"inject_cover|cover"],
     description="Conversion Markdown -> DOCX avec pandoc + inject_cover.py")

test("D08", "agent_docx_complete", "DOCX",
     "Convertis ce texte en PDF professionnel avec mise en forme corporate :\n\n"
     "Titre : Rapport mensuel\nAuteur : Damien\n\n"
     "Section 1 : Résumé exécutif\nLe mois d'avril a été marqué par une croissance de 15%.\n\n"
     "Section 2 : Détails\n- Ventes : 150k€\n- Charges : 120k€\n- Résultat : +30k€",
     patterns=[r"pdf|soffice|convert"],
     description="Conversion DOCX -> PDF via soffice",
     expect_file=True, file_ext=".pdf")

test("D10", "agent_docx_complete", "DOCX",
     "Crée un document Word 'Fiche produit' avec : titre 'Widget Pro X200', "
     "un tableau de spécifications (Poids: 1.2kg, Dimensions: 30x20x10cm, Prix HT: 149.90€), "
     "et un paragraphe de description marketing.",
     patterns=[
         r"fill_template|template.*corporate|template.*base",
         r"execute_code",
     ],
     description="Création avec fill_template.py + type table",
     expect_file=True, file_ext=".docx")

test("D13", "agent_docx_complete", "DOCX",
     "Rédige un courrier professionnel adressé à M. Pascal Caiozzo, Arémont SAS, 75008 Paris. "
     "Objet : proposition d'accompagnement IA pour les équipes achats. "
     "Contenu : suite à notre échange, nous proposons une intervention en 3 étapes (keynote, formation, acculturation) "
     "pour un budget de 11 900€ HT. Signé par Jean Dupont, Consultant IA, Acme Corp.",
     patterns=[
         r"fill_courrier|template.*courrier|courrier",
         r"execute_code",
     ],
     description="Créer un courrier via fill_courrier_template.py",
     expect_file=True, file_ext=".docx")

test("D14", "agent_docx_complete", "DOCX",
     "Fais un document Word à partir de ce markdown de réunion :\n\n"
     "# Compte-rendu de réunion - Projet Alpha\n\n"
     "## Informations\n\n- **Date :** 15 avril 2026\n- **Lieu :** Visioconférence\n\n"
     "## Participants\n\n- Sophie Martin (DRH, Nextera)\n- Jean Dupont (Consultant, Acme)\n\n"
     "## Décisions\n\n1. **Lancement POC** prévu le 1er mai\n   - Budget : 15k€\n   - Pilote : équipe RH\n"
     "2. **Formation** des managers en juin\n   - 3 sessions de 2h\n\n"
     "## Actions\n\n- Damien : envoyer proposition avant le 20 avril\n- Sophie : valider budget interne",
     patterns=[
         r"fill_cr_template|template.*CR|compte.rendu",
         r"execute_code",
     ],
     description="Markdown de type CR → doit choisir fill_cr_template (PAS pandoc)",
     expect_file=True, file_ext=".docx")

test("D15", "agent_docx_complete", "DOCX",
     "Voici un compte-rendu en markdown. Fais-en un Word :\n\n"
     "# CR Réunion Projet Beta\n\n## Infos\n\n- **Date :** 18/04/2026\n- **Lieu :** Teams\n\n"
     "## Participants\n\n- Alice Dupont (Chef de projet, ClientCo)\n- Bob Martin (Dev Lead, Acme)\n\n"
     "## Points abordés\n\n1. **Planning sprint 4**\n   - Livraison prévue le 30 avril\n"
     "   - 3 stories restantes\n2. **Bug critique #142**\n   - Corrigé en prod\n   - Post-mortem prévu\n\n"
     "## Actions\n\n- Alice : valider la recette avant le 25\n- Bob : préparer le post-mortem",
     patterns=[
         r"fill_cr_template|template.*CR|compte.rendu",
         r"subitems|sub.?items",
         r"execute_code",
     ],
     description="Markdown CR avec listes imbriquées → fill_cr_template avec subitems",
     expect_file=True, file_ext=".docx")


# ==========================================
# DOCX Agent — Tests nécessitant un fichier source
# ==========================================

test("D01", "agent_docx_complete", "DOCX",
     "Voici un rapport d'audit DAF en français (D01_anonymized.docx) : 6 pages avec tableaux, "
     "images et pieds de page. Analyse sa structure et sa mise en forme (styles, polices, couleurs, "
     "en-têtes, pieds de page). Puis produis un nouveau CR dans un format strictement identique : "
     "Réunion produit du 14 avril 2026. Participants : Marie, Jean, Sophie. "
     "Décision : lancement V2 le 15 juin. Action : Jean prépare le plan média avant le 1er mai.",
     patterns=[r"unpack|lxml|template|analyse"],
     description="Reproduire un CR depuis template utilisateur",
     expect_file=True, file_ext=".docx")

test("D03", "agent_docx_complete", "DOCX",
     "Voici un fichier de CGV en anglais (D02.docx, ~11 pages). D'abord, remplace toutes les "
     "occurrences de 'the Client' par 'the User' en tracked changes. Puis accepte tous les "
     "tracked changes et produis la version finale propre du document.",
     patterns=[r"accept_changes|accept|tracked"],
     description="Tracked changes + accepter -> version propre",
     expect_file=True, file_ext=".docx")

test("D04", "agent_docx_complete", "DOCX",
     "Voici une proposition commerciale de formation IA (D04_test_03_proposition_commerciale_variantes.docx, "
     "1 page en français). Relis-la et ajoute les commentaires Word suivants : "
     "(1) sur le premier paragraphe décrivant les prestations, commente "
     "'À détailler avec les cas d'usage spécifiques du client', "
     "(2) sur le montant ou tarif mentionné, commente "
     "'Remise de 10% à négocier pour contrat annuel ?'.",
     patterns=[r"comment|commentaire|unpack"],
     description="Ajouter commentaires de relecture",
     expect_file=True, file_ext=".docx")

test("D06", "agent_docx_complete", "DOCX",
     "Voici deux documents Word à fusionner : D06_a_anonymized.docx (proposition technique, "
     "3 pages avec tableaux et images) et D06_b_anonymized.docx (conditions d'exécution, 2 pages). "
     "Fusionne-les en un seul document : la proposition technique en premier, un saut de page, "
     "puis les conditions d'exécution. Conserve la mise en forme de chaque document.",
     patterns=[r"fusionne|merge|docxcompose|page_break"],
     description="Fusion de deux DOCX",
     expect_file=True, file_ext=".docx")

test("D07", "agent_docx_complete", "DOCX",
     "Voici un rapport de rémunération des dirigeants FY2025 en anglais (D07.docx, ~13 pages, "
     "37 tableaux, 5 images). Analyse ce document Word et extrais-moi un résumé structuré : "
     "liste des titres de sections, nombre de tableaux, nombre d'images, et le texte des "
     "3 premiers paragraphes.",
     patterns=[r"unpack|pandoc|python-docx|analyse|section"],
     description="Extraction contenu structuré DOCX complexe")

test("D09", "agent_docx_complete", "DOCX",
     "Ce fichier (D09.doc, 186 Ko) est dans l'ancien format Word .doc binaire. "
     "Convertis-le en .docx moderne en conservant toute la mise en forme et le contenu.",
     patterns=[r"soffice|convert.*docx"],
     description="Conversion .doc legacy -> .docx",
     expect_file=True, file_ext=".docx")

test("D11", "agent_docx_complete", "DOCX",
     "Voici un contrat de travail CDD en français (D11.docx, ~6 pages). Dans ce document, "
     "seule la première occurrence de 'Directeur' doit être remplacée par 'Directrice' "
     "(c'est un changement de titre pour la DG uniquement). Fais-le en tracked changes.",
     patterns=[r"tracked.replace|tracked.change|--first|first|première"],
     description="Remplacement ciblé --first en tracked changes",
     expect_file=True, file_ext=".docx")

test("D12", "agent_docx_complete", "DOCX",
     "Voici un formulaire d'autorisation en anglais (D12.docx, 9 pages, 23 tableaux, 8 sections). "
     "Remplis-le avec : consultant=Marie Dupont, client=Société ABC, date=1er mai 2026, "
     "durée=6 mois, tarif=850€ HT. Puis remplace 'the terms defined' par "
     "'the revised terms of the framework agreement' en tracked changes. "
     "Exporte le résultat final en PDF.",
     patterns=[r"unpack|tracked|soffice|pdf"],
     description="Pipeline complet template -> tracked changes -> PDF",
     expect_file=True, file_ext=".pdf")


# ==========================================
# PPTX Agent — Self-contained tests
# ==========================================

test("P01b", "agent_pptx_complete", "PPTX",
     "Crée une présentation de 5 slides sur l'IA générative pour une réunion interne.",
     patterns=[
         r"pptxgenjs|pptxgen|PptxGenJS|create_from_template|template.*corporate|template.*corporate",
         r"execute_code",
     ],
     description="Création PPTX avec template corporate ou pptxgenjs",
     expect_file=True, file_ext=".pptx")

test("P01", "agent_pptx_complete", "PPTX",
     "Crée un pitch deck de 6 slides pour une startup EdTech appelée 'LearnAI'. "
     "Slides : Titre, Problème, Solution, Marché (TAM 50Md€), Business model, Ask (levée 1.5M€). "
     "Design moderne, palette verte/blanche.",
     patterns=[
         r"pptxgenjs|pptxgen|PptxGenJS|Node",
         r"execute_code",
     ],
     description="Création pitch deck pptxgenjs",
     expect_file=True, file_ext=".pptx")

test("P08", "agent_pptx_complete", "PPTX",
     "Crée une présentation de 3 slides avec des graphiques : "
     "(1) Titre 'Résultats T1', "
     "(2) Barres ventes par région (Nord:45k, Sud:38k, Est:52k, Ouest:41k), "
     "(3) Camembert charges (Salaires:60%, Loyer:15%, Marketing:20%, Divers:5%).",
     patterns=[
         r"pptxgenjs|pptxgen|chart|graphique",
         r"execute_code",
     ],
     description="Création slides avec graphiques pptxgenjs",
     expect_file=True, file_ext=".pptx")

test("P11", "agent_pptx_complete", "PPTX",
     "Crée une mini-présentation de 3 slides : (1) Titre 'Résultats T1', "
     "(2) Barres ventes par région, (3) Camembert répartition charges.",
     patterns=[r"pptxgenjs|pptxgen|chart|addChart"],
     description="Création PPTX avec graphiques",
     expect_file=True, file_ext=".pptx")


# ==========================================
# PPTX Agent — Tests nécessitant un fichier source
# ==========================================

test("P02", "agent_pptx_complete", "PPTX",
     "Voici un deck 'Researcher Deep Dive' sur l'IA (P02.pptx, 20 slides, 9 Mo). "
     "Remplace le titre du slide 1 par 'Bilan annuel 2025' et le sous-titre par "
     "'Direction Commerciale'. Mets à jour la date en pied de page sur tous les slides "
     "à 'Avril 2026'. Conserve strictement la charte graphique existante.",
     patterns=[r"unpack|xml|edit|slide"],
     description="Edition template PPTX existant",
     expect_file=True, file_ext=".pptx")

test("P03", "agent_pptx_complete", "PPTX",
     "Voici un template PowerPoint (P03.pptx, 20 slides). Analyse-le : montre-moi un aperçu "
     "visuel (thumbnails) de tous les layouts disponibles et dis-moi quels types de slides "
     "je peux créer avec chacun.",
     patterns=[r"thumbnail|layout|aperçu|analyse"],
     description="Analyse template avec thumbnails")

test("P04", "agent_pptx_complete", "PPTX",
     "Voici une présentation (P04.pptx) avec un slide daté du 30/06/2025. "
     "Duplique ce slide 3 fois pour avoir 4 copies au total.",
     patterns=[r"add_slide|dupli|copie"],
     description="Duplication de slides",
     expect_file=True, file_ext=".pptx")

test("P05", "agent_pptx_complete", "PPTX",
     "Ce fichier PowerPoint (P05.pptx) fait 19 Mo avec 49 slides (deck corporate "
     "'Conseil et Opérateur d'IA'). C'est trop lourd pour l'envoi par email. "
     "Nettoie-le : supprime les slides masqués, les médias non référencés, et optimise la taille.",
     patterns=[r"clean|nettoie|orphan"],
     description="Nettoyage PPTX volumineux 19 Mo",
     expect_file=True, file_ext=".pptx")

test("P06", "agent_pptx_complete", "PPTX",
     "Convertis cette présentation (P02.pptx, 20 slides) en PDF fidèle pour diffusion "
     "aux participants de la réunion.",
     patterns=[r"soffice|pdf|convert"],
     description="Conversion PPTX -> PDF",
     expect_file=True, file_ext=".pdf")

test("P07", "agent_pptx_complete", "PPTX",
     "J'ai besoin du contenu textuel de cette présentation corporate (P05.pptx, 49 slides, "
     "'Conseil et Opérateur d'IA'). Extrais tout le texte slide par slide en markdown structuré.",
     patterns=[r"markitdown|markdown|extract"],
     description="Extraction contenu PPTX en markdown")

test("P09", "agent_pptx_complete", "PPTX",
     "Voici un template PowerPoint vide Acme Corp (P09.pptx) avec 47 layouts disponibles "
     "mais aucun slide. Ajoute un nouveau slide en utilisant le 2ème layout disponible. "
     "Titre : 'Prochaines étapes', contenu : 'Valider le budget', 'Recruter 2 développeurs', "
     "'Lancer la V2 en juin'.",
     patterns=[r"add_slide|layout|unpack"],
     description="Ajout slide depuis layout sur template vide",
     expect_file=True, file_ext=".pptx")

test("P10", "agent_pptx_complete", "PPTX",
     "Voici un deck corporate en français (P10.pptx, 26 slides, 5 Mo). "
     "Remplace toutes les occurrences de 'Acme Corp' par 'corporate Consulting' "
     "dans tous les slides, y compris les masters et layouts.",
     patterns=[r"replace|remplac|unpack|xml"],
     description="Remplacement texte dans toute la présentation",
     expect_file=True, file_ext=".pptx")

test("P12", "agent_pptx_complete", "PPTX",
     "Voici un template PowerPoint (.potx) avec 18 layouts et 8 slides de référence (P12.potx). "
     "Analyse ce template corporate, puis crée 3 nouveaux slides dans le même style : "
     "'Objectifs 2026' avec 4 bullet points, 'Budget prévisionnel' avec un tableau 4x3, "
     "et 'Calendrier'. Exporte le résultat en PDF.",
     patterns=[r"thumbnail|analyse|add_slide|soffice|pdf"],
     description="Pipeline analyse template -> création -> export PDF",
     expect_file=True, file_ext=".pdf")


# ==========================================
# XLSX Agent — Self-contained tests
# ==========================================

test("X01", "agent_xlsx_complete", "XLSX",
     "Crée un fichier Excel de budget trimestriel avec un onglet Q1 contenant les postes "
     "Salaires (45000€), Loyer (8000€), Marketing (12000€), IT (6000€), et un sous-total en formule Excel. "
     "Formate avec en-têtes colorés et format monétaire €.",
     patterns=[
         r"openpyxl|Excel|xlsx",
         r"execute_code",
     ],
     description="Création Excel avec openpyxl + styles",
     expect_file=True, file_ext=".xlsx")

test("X03", "agent_xlsx_complete", "XLSX",
     "Crée un fichier Excel avec une colonne Prix (10, 20, 30) et une colonne Total qui fait =Prix*1.2. "
     "Recalcule les formules pour que les valeurs soient visibles.",
     patterns=[
         r"openpyxl|formul|recalc",
         r"execute_code",
     ],
     description="Création Excel + recalc.py",
     expect_file=True, file_ext=".xlsx")

test("X07", "agent_xlsx_complete", "XLSX",
     "Crée un fichier Excel de suivi des objectifs avec 10 vendeurs : vert si résultat >= objectif, "
     "orange entre 80-100%, rouge si < 80%.",
     patterns=[r"openpyxl|Conditional|mise.en.forme|couleur"],
     description="Mise en forme conditionnelle",
     expect_file=True, file_ext=".xlsx")

test("X11", "agent_xlsx_complete", "XLSX",
     "Crée un modèle de prévision de trésorerie sur 12 mois : solde initial 50k€, "
     "encaissements croissants +5%/mois depuis 20k€, décaissements fixes 18k€. "
     "Formules Excel natives, négatifs en rouge.",
     patterns=[r"openpyxl|formul|=SUM|trésorerie|prévision"],
     description="Modèle financier avec formules complexes",
     expect_file=True, file_ext=".xlsx")


# ==========================================
# XLSX Agent — Tests nécessitant un fichier source
# ==========================================

test("X02", "agent_xlsx_complete", "XLSX",
     "Voici un export CRM de données commerciales (X02_donnees_commerciales_complexes.xlsx) : "
     "3000 commandes sur 28 colonnes, avec montants HT/TVA/TTC et multi-onglets (commandes, "
     "clients, produits). Analyse ce fichier : (1) nombre de lignes/colonnes par onglet, "
     "(2) types de données par colonne, (3) valeurs manquantes, (4) top 5 clients par CA, "
     "(5) répartition mensuelle du CA.",
     patterns=[r"pandas|read_excel|describe|info|groupby"],
     description="Analyse Excel commercial 3000 lignes")

test("X04", "agent_xlsx_complete", "XLSX",
     "Voici un fichier de ventes mensuelles (X04_ventes_mensuelles_input_complexe.xlsx) avec "
     "24 mois de données par canal (en ligne, boutique, etc.) incluant budget et variance. "
     "Ajoute un graphique en barres montrant l'évolution des ventes mensuelles totales. "
     "Place le graphique dans un nouvel onglet 'Dashboard'.",
     patterns=[r"openpyxl|BarChart|chart|Dashboard"],
     description="Graphique barres dans Excel 24 mois",
     expect_file=True, file_ext=".xlsx")

test("X05", "agent_xlsx_complete", "XLSX",
     "Ce fichier (X05_ancien_format_complexe.xls) est dans l'ancien format Excel .xls "
     "avec 2 onglets annuels et des formules legacy. Convertis-le en .xlsx moderne "
     "sans perdre les données ni les formules.",
     patterns=[r"soffice|convert.*xlsx"],
     description="Conversion XLS legacy -> XLSX",
     expect_file=True, file_ext=".xlsx")

test("X06", "agent_xlsx_complete", "XLSX",
     "Voici un fichier de 5000 transactions commerciales "
     "(X06_transactions_pivot_input_complexe.xlsx) avec catégories, sous-catégories et dates. "
     "Crée un tableau croisé dynamique montrant le CA par catégorie de produit et par mois, "
     "dans un nouvel onglet 'Pivot'.",
     patterns=[r"pivot|crois|pandas"],
     description="Tableau croisé dynamique sur 5000 transactions",
     expect_file=True, file_ext=".xlsx")

test("X08", "agent_xlsx_complete", "XLSX",
     "Voici un rapport de type board pack (X08_source_export_pdf_complexe.xlsx) avec "
     "17 colonnes de données financières (budget, réalisé, variance, YTD). "
     "Exporte-le en PDF en mode paysage, adapté à la largeur d'une page A4.",
     patterns=[r"soffice|pdf|convert|paysage"],
     description="Export Excel board pack -> PDF paysage",
     expect_file=True, file_ext=".pdf")

test("X09", "agent_xlsx_complete", "XLSX",
     "Voici 3 rapports mensuels P&L (X09_rapport_complexe_janvier.xlsx, "
     "X09_rapport_complexe_fevrier.xlsx, X09_rapport_complexe_mars.xlsx), chacun avec "
     "un P&L et un détail clients sur plusieurs onglets. Fusionne-les en un seul fichier "
     "avec un onglet par mois source et un onglet 'Consolidé' qui additionne les valeurs.",
     patterns=[r"pandas|openpyxl|fusionne|merge|consolid"],
     description="Fusion de 3 Excel P&L mensuels",
     expect_file=True, file_ext=".xlsx")

test("X10", "agent_xlsx_complete", "XLSX",
     "Voici un export CRM sale (X10_donnees_sales_complexes.xlsx, 332 lignes) avec des "
     "problèmes de qualité : 8 doublons exacts, formats de dates mixtes, casse incohérente "
     "dans les noms, emails et téléphones incomplets. Nettoie ce fichier : supprime les "
     "doublons, normalise les noms (casse), corrige les dates, et produis un rapport des "
     "modifications effectuées.",
     patterns=[r"pandas|drop_duplicates|netto|clean"],
     description="Nettoyage et dédoublonnage CRM 332 lignes",
     expect_file=True, file_ext=".xlsx")

test("X12", "agent_xlsx_complete", "XLSX",
     "Voici un classeur RH complet (X12_donnees_rh_complexes.xlsx) avec 4 onglets : "
     "850 collaborateurs, 10200 lignes de paie, et 1200 absences. Fais une analyse complète : "
     "effectif et masse salariale par département, salaire moyen/médian, et crée des graphiques "
     "Excel (barres + camembert) dans un onglet 'Dashboard RH'.",
     patterns=[r"pandas|openpyxl|chart|Dashboard|analyse"],
     description="Analyse RH 850 employés + Dashboard",
     expect_file=True, file_ext=".xlsx")


# ==========================================
# PDF Agent — Self-contained tests
# ==========================================

test("F13", "agent_pdf_complete", "PDF",
     "Crée un PDF professionnel 'Proposition commerciale' avec 3 sections : "
     "Contexte, Offre de service (avec liste à puces), et Conditions tarifaires.",
     patterns=[
         r"pdf|PDF|soffice|template|fill_template",
         r"execute_code",
     ],
     description="Création PDF via DOCX corporate -> soffice",
     expect_file=True, file_ext=".pdf")

test("F04", "agent_pdf_complete", "PDF",
     "Crée deux PDFs simples (une page chacun avec du texte) puis fusionne-les en un seul.",
     patterns=[
         r"fusion|merge|fusionne|PdfMerger|qpdf",
         r"execute_code",
     ],
     description="Fusion de PDFs",
     expect_file=True, file_ext=".pdf")


# ==========================================
# PDF Agent — Tests nécessitant un fichier source
# ==========================================

test("F01", "agent_pdf_complete", "PDF",
     "Voici un contrat de service en anglais (service-agreement.pdf, 15 pages). "
     "Extrais le texte et identifie les clauses principales : durée, montant, "
     "conditions de résiliation, pénalités. Présente un résumé structuré.",
     patterns=[r"pdfplumber|extract|clause|texte"],
     description="Extraction texte contrat PDF 15 pages")

test("F02", "agent_pdf_complete", "PDF",
     "Voici une facture en PDF (invoice-sample.pdf) contenant des tableaux de lignes "
     "de facturation avec montants et descriptions. Extrais les tableaux et convertis-les "
     "en fichier Excel exploitable.",
     patterns=[r"pdfplumber|extract_table|pandas|Excel"],
     description="Extraction tableaux facture PDF -> Excel",
     expect_file=True, file_ext=".xlsx")

test("F03", "agent_pdf_complete", "PDF",
     "Ce PDF (scan-sample.pdf) est un document scanné (image, pas de texte extractible). "
     "Extrais le texte par OCR et produis un fichier texte lisible.",
     patterns=[r"pdf2image|pytesseract|OCR|image_to_string"],
     description="OCR sur PDF scanné")

test("F05", "agent_pdf_complete", "PDF",
     "Voici un document de 15 pages (document-15pages.pdf). Extrais les pages 3 à 7 "
     "dans un fichier séparé 'extrait.pdf'.",
     patterns=[r"pypdf|PdfReader|PdfWriter|qpdf|pages"],
     description="Split PDF 15 pages -> extraction pages 3-7",
     expect_file=True, file_ext=".pdf")

test("F06", "agent_pdf_complete", "PDF",
     "Voici 3 fichiers PDF à vérifier : corrupted.pdf (1 Ko, probablement corrompu), "
     "encrypted.pdf (possiblement chiffré), et not_encrypted.pdf (référence normale). "
     "Vérifie l'intégrité de chacun et répare ceux qui le nécessitent. "
     "Donne un rapport d'état par fichier.",
     patterns=[r"qpdf|check|repair|intégrité"],
     description="Vérification intégrité et réparation 3 PDFs")

test("F07", "agent_pdf_complete", "PDF",
     "Voici un rapport de 16 pages (document-a-convertir.pdf). Convertis chaque page "
     "en image PNG haute résolution (300 DPI).",
     patterns=[r"pdf2image|pdftoppm|convert.*image|300|dpi"],
     description="Conversion PDF 16 pages -> images PNG 300 DPI",
     expect_file=True, file_ext=".png")

test("F08", "agent_pdf_complete", "PDF",
     "Voici une facture PDF (invoice-metadata.pdf, 1 page). Donne-moi toutes les "
     "métadonnées : auteur, date de création, date de modification, producteur, "
     "nombre de pages, taille, version PDF, et s'il est chiffré ou non.",
     patterns=[r"pypdf|PdfReader|metadata|qpdf"],
     description="Extraction métadonnées PDF facture")

test("F09", "agent_pdf_complete", "PDF",
     "Ce document PDF (document-pages-retournees.pdf, 15 pages) a certaines pages "
     "retournées à 180°. Identifie les pages mal orientées et corrige-les.",
     patterns=[r"pypdf|rotate|rotation"],
     description="Correction rotation pages PDF")

test("F10", "agent_pdf_complete", "PDF",
     "Voici un document de 15 pages (document-a-watermarker.pdf). Ajoute un watermark "
     "'BROUILLON' en diagonale sur toutes les pages, en texte gris semi-transparent.",
     patterns=[r"reportlab|watermark|merge|Canvas"],
     description="Ajout watermark BROUILLON sur 15 pages",
     expect_file=True, file_ext=".pdf")

test("F11", "agent_pdf_complete", "PDF",
     "Ce PDF (sample-heavy-25mb.pdf) fait plus de 100 Mo. Optimise-le pour réduire "
     "significativement sa taille tout en conservant le contenu lisible.",
     patterns=[r"qpdf|compress|optimis|linearize"],
     description="Compression PDF lourd 100+ Mo",
     expect_file=True, file_ext=".pdf")

test("F12", "agent_pdf_complete", "PDF",
     "Ce relevé bancaire (bank-statement-scanned.pdf) est un PDF scanné. Extrais les "
     "transactions (date, libellé, débit, crédit) par OCR et produis un fichier Excel structuré.",
     patterns=[r"pdf2image|pytesseract|pandas|OCR|Excel"],
     description="Pipeline OCR releve bancaire -> Excel",
     expect_file=True, file_ext=".xlsx")


# ==========================================
# FFmpeg Agent — Self-contained tests
# ==========================================

test("M05", "agent_quick_edits", "FFmpeg",
     "Analyse les capacités de ffmpeg installé : quels codecs vidéo et audio sont disponibles ? "
     "Liste les 5 principaux codecs vidéo et audio.",
     patterns=[r"ffmpeg|ffprobe|codec"],
     description="Analyse ffmpeg capabilities")

test("M06", "agent_quick_edits", "FFmpeg",
     "Crée une image PNG de 800x600 pixels, fond bleu (#2F5597), avec le texte 'Test' en blanc centré.",
     patterns=[r"PIL|Pillow|image|Image|png|PNG"],
     description="Création d'image avec Pillow",
     expect_file=True, file_ext=".png")


# ==========================================
# FFmpeg Agent — Tests nécessitant un fichier source
# ==========================================

test("M01", "agent_quick_edits", "FFmpeg",
     "Voici une vidéo au format MOV (file_example_MOV_640_800kB.mov, 640px, 778 Ko). "
     "Convertis-la en MP4 compatible web (H.264 + AAC), résolution 720p max si nécessaire.",
     patterns=[r"ffmpeg|ffprobe|libx264|aac|720"],
     description="Conversion MOV 640px -> MP4 H.264",
     expect_file=True, file_ext=".mp4")

test("M02", "agent_quick_edits", "FFmpeg",
     "Voici une vidéo MP4 de 10 secondes (sample-10s.mp4, 5 Mo). "
     "Extrais uniquement la piste audio en MP3 à 192 kbps.",
     patterns=[r"ffmpeg|-vn|mp3lame|audio"],
     description="Extraction audio MP3 depuis vidéo 10s",
     expect_file=True, file_ext=".mp3")

test("M03", "agent_quick_edits", "FFmpeg",
     "Voici une vidéo de surf en 720p (sample_1280x720_surfing_with_audio.avi, "
     "AVI, ~3 minutes, 26 Mo). Découpe-la pour garder uniquement la partie de 0:45 à 2:30.",
     patterns=[r"ffmpeg|-ss|-to|cut|découpe"],
     description="Découpe AVI surf 0:45 -> 2:30",
     expect_file=True, file_ext=".mp4")

test("M04", "agent_quick_edits", "FFmpeg",
     "Voici une vidéo de 10 secondes (sample-10s.mp4). "
     "Crée un GIF animé à partir des 5 premières secondes, 320px de large, 10 fps.",
     patterns=[r"ffmpeg|gif|fps|scale"],
     description="Création GIF depuis vidéo 10s",
     expect_file=True, file_ext=".gif")

test("M07", "agent_quick_edits", "FFmpeg",
     "Voici une photo de paysage montagneux (Landscape_big_river_in_mountains.jpg, "
     "1600x1066 px, 194 Ko). Ajoute en bas un bandeau noir semi-transparent avec "
     "le texte '© example.com 2026' en blanc, centré.",
     patterns=[r"PIL|Pillow|ImageDraw|texte|bandeau"],
     description="Ajout bandeau texte sur photo paysage",
     expect_file=True, file_ext=".jpg")

test("M08", "agent_quick_edits", "FFmpeg",
     "Voici deux vidéos MP4 : sample-15s.mp4 (15 secondes) et sample-20s.mp4 "
     "(20 secondes). Assemble-les bout à bout dans cet ordre.",
     patterns=[r"ffmpeg|concat|assemble|fusion"],
     description="Concaténation 2 vidéos MP4",
     expect_file=True, file_ext=".mp4")

test("M09", "agent_quick_edits", "FFmpeg",
     "Voici une vidéo (sample-10s.mp4, 10 secondes) et un fichier audio "
     "(sample-12s.mp3, 12 secondes). Ajoute la musique en fond sonore à la vidéo, "
     "volume musique à 20% par rapport à l'audio original.",
     patterns=[r"ffmpeg|amix|amerge|volume|audio"],
     description="Mixage audio 20% sur vidéo",
     expect_file=True, file_ext=".mp4")

test("M10", "agent_quick_edits", "FFmpeg",
     "Voici 4 images de 400x300 pixels : sample-boat-400x300.png, "
     "sample-city-park-400x300.jpg, sample-clouds-400x300.jpg et "
     "sample-birch-400x300.jpg. Crée une mosaïque 2x2 en les assemblant.",
     patterns=[r"PIL|Pillow|paste|mosaïque|xstack"],
     description="Mosaique 2x2 de 4 images 400x300",
     expect_file=True, file_ext=".png")

test("M11", "agent_quick_edits", "FFmpeg",
     "Voici une vidéo de surf en 720p (sample_1280x720_surfing_with_audio.avi, ~3 min). "
     "Extrais une capture d'écran à 1 minute et 23 secondes, en PNG pleine résolution.",
     patterns=[r"ffmpeg|-ss|frames|capture|screenshot"],
     description="Extraction frame a 1:23 depuis AVI surf",
     expect_file=True, file_ext=".png")

test("M12", "agent_quick_edits", "FFmpeg",
     "Voici 5 images en formats variés : sample-boat-400x300.png, "
     "sample-city-park-400x300.jpg, sample-clouds-400x300.jpg, "
     "sample-clouds2-400x300.png et sample-bumblebee-400x300.png. "
     "Convertis-les toutes en JPEG, 1024px de large max (sans agrandir les plus petites), "
     "qualité 90%. Renomme-les photo_01.jpg à photo_05.jpg.",
     patterns=[r"PIL|Pillow|thumbnail|JPEG|convert|batch"],
     description="Conversion batch 5 images -> JPEG 1024px",
     expect_file=True, file_ext=".jpg")


# ==========================================
# DataViz Agent — Self-contained tests
# ==========================================

test("A01", "agent_data_viz", "DataViz",
     "Génère un dataset fictif de 100 ventes (date, produit parmi A/B/C, montant entre 50 et 500€, "
     "région parmi Nord/Sud/Est/Ouest). Puis fais une analyse exploratoire : "
     "statistiques descriptives, répartition par produit, et un histogramme des montants.",
     patterns=[
         r"pandas|DataFrame|analyse|exploratoire|dataset|vente",
     ],
     description="Analyse exploratoire + visualisation",
     expect_file=True, file_ext=".png")

test("A02", "agent_data_viz", "DataViz",
     "Crée un dashboard en une seule image avec 4 graphiques à partir de données fictives : "
     "(1) courbe CA mensuel, (2) camembert par catégorie, (3) barres top clients, "
     "(4) scatter quantité vs montant. Utilise la palette de couleurs corporate.",
     patterns=[
         r"dashboard|subplots|graphique|corporate|palette|4.*graph",
     ],
     description="Dashboard 4 graphiques avec palette corporate",
     expect_file=True, file_ext=".png")

test("A03", "agent_data_viz", "DataViz",
     "Génère des données de satisfaction client par canal (boutique, web, téléphone, 50 par canal). "
     "Fais un test ANOVA pour savoir si la satisfaction diffère entre canaux. "
     "Donne F-stat, p-value et un boxplot.",
     patterns=[r"ANOVA|f_oneway|scipy|boxplot|satisfaction"],
     description="Test statistique ANOVA",
     expect_file=True, file_ext=".png")

test("A04", "agent_data_viz", "DataViz",
     "Génère un dataset avec 5 variables numériques corrélées (200 lignes). "
     "Calcule la matrice de corrélation et affiche-la en heatmap annotée. "
     "Identifie les corrélations > 0.7.",
     patterns=[r"corr|heatmap|corrélation|seaborn"],
     description="Corrélation et heatmap",
     expect_file=True, file_ext=".png")

test("A05", "agent_data_viz", "DataViz",
     "Génère 200 points avec relation linéaire bruitée surface_m2 (30-150) vs prix (3.5k€/m²). "
     "Entraîne une régression linéaire, affiche R², équation, droite + intervalle de confiance.",
     patterns=[r"LinearRegression|regression|sklearn|R2|scatter"],
     description="Régression linéaire avec prédiction",
     expect_file=True, file_ext=".png")

test("A06", "agent_data_viz", "DataViz",
     "Génère des données clients (âge, revenu, dépenses, 300 lignes). "
     "Segmente avec K-Means, méthode du coude, visualise en 2D avec PCA.",
     patterns=[r"KMeans|cluster|PCA|coude|elbow"],
     description="Clustering K-Means",
     expect_file=True, file_ext=".png")

test("A07", "agent_data_viz", "DataViz",
     "Génère une série temporelle de CA quotidien sur 1 an (tendance haussière + saisonnalité). "
     "Fais une décomposition et une prévision naïve sur 30 jours.",
     patterns=[r"seasonal_decompose|statsmodels|série.temporelle|prévision"],
     description="Analyse séries temporelles",
     expect_file=True, file_ext=".png")

test("A08", "agent_data_viz", "DataViz",
     "Génère un dataset de ventes (100 lignes, colonnes: date, produit, région, montant). "
     "Exporte dans un Excel 3 onglets : données brutes, statistiques par catégorie, tableau croisé.",
     patterns=[r"ExcelWriter|openpyxl|onglet|sheet|pivot"],
     description="Export multi-onglets vers Excel",
     expect_file=True, file_ext=".xlsx")

test("A09", "agent_data_viz", "DataViz",
     "Génère 200 transactions avec 5% d'anomalies (montants aberrants). "
     "Détecte les outliers avec IQR et Z-score. Visualise sur un boxplot.",
     patterns=[r"zscore|IQR|outlier|anomalie|boxplot"],
     description="Détection d'anomalies",
     expect_file=True, file_ext=".png")

test("A10", "agent_data_viz", "DataViz",
     "Génère deux échantillons : 'avant' (n=50, μ=72, σ=12) et 'après' (n=50, μ=78, σ=11). "
     "Fais un test t de Student apparié, donne la conclusion, visualise les distributions.",
     patterns=[r"ttest|Student|scipy|distribution|avant.*après"],
     description="Comparaison avant/après avec test t",
     expect_file=True, file_ext=".png")

test("A11", "agent_data_viz", "DataViz",
     "Génère un nuage de mots à partir de ce texte : "
     "'L intelligence artificielle transforme les entreprises. Le machine learning et le deep learning "
     "permettent d automatiser des tâches complexes. Les données sont le carburant de l IA. "
     "Les algorithmes apprennent des patterns dans les données pour prédire et classifier.' "
     "Exclue les mots vides français, palette bleue.",
     patterns=[r"wordcloud|WordCloud|nuage|cloud"],
     description="Word cloud",
     expect_file=True, file_ext=".png")

test("A12", "agent_data_viz", "DataViz",
     "Génère un CSV de données sales (100 lignes, doublons, valeurs manquantes, types incohérents). "
     "Pipeline complet : charge, nettoie, analyse, 3 visualisations, export Excel avec images intégrées.",
     patterns=[r"pandas|matplotlib|openpyxl|pipeline|nettoie|clean"],
     description="Pipeline complet données -> analyse -> viz -> export",
     expect_file=True, file_ext=".xlsx")


# === Test Runner ===

def upload_files_for_test(test_case: TestCase) -> list:
    """Upload input files for a test. Returns list of file_ids (may be empty)."""
    file_ids = []
    file_rel_paths = TEST_INPUT_FILES.get(test_case.test_id, [])
    if not file_rel_paths:
        return file_ids

    for rel_path in file_rel_paths:
        file_path = TEST_FILES_DIR / rel_path
        if not file_path.exists():
            print(f"    WARNING: Input file missing: {file_path}")
            continue
        file_id = upload_file(file_path)
        if file_id:
            file_ids.append(file_id)
            print(f"    Uploaded: {file_path.name} -> {file_id[:16]}...")
        else:
            print(f"    INFO: File upload skipped (no JWT): {file_path.name}")

    return file_ids


def run_test(test_case: TestCase) -> TestResult:
    """Run a single test and return the result."""
    result = TestResult(
        test_id=test_case.test_id,
        agent=test_case.agent_name,
        prompt=test_case.prompt[:100] + "...",
    )

    start = time.time()
    try:
        # Upload files if needed
        file_ids = upload_files_for_test(test_case)
        result.files_uploaded = file_ids

        response = call_agent(
            test_case.agent_id,
            test_case.prompt,
            file_ids=file_ids if file_ids else None,
            previous_response_id=test_case.previous_response_id,
        )
        result.duration = time.time() - start

        if response.get("status") != "completed":
            result.status = "ERROR"
            result.error = f"Response status: {response.get('status')}"
            return result

        if response.get("error"):
            result.status = "ERROR"
            result.error = str(response["error"])
            return result

        text, code_blocks, files = extract_response_data(response)
        result.response_text = text[:500]
        result.code_blocks = code_blocks
        result.files_generated = files

        # Check methodology
        result.methodology_checks = check_methodology(
            code_blocks,
            test_case.methodology_patterns,
            test_case.methodology_antipatterns,
        )

        # Determine pass/fail
        has_missing = any(v == "MISSING" for v in result.methodology_checks.values())
        has_violation = any(v == "VIOLATION" for v in result.methodology_checks.values())

        if has_violation:
            result.status = "FAIL"
            result.error = "Methodology antipattern detected"
        elif has_missing:
            result.status = "FAIL"
            result.error = "Expected methodology pattern not found"
        else:
            result.status = "PASS"

    except requests.exceptions.Timeout:
        result.status = "ERROR"
        result.error = "Timeout"
        result.duration = time.time() - start
    except Exception as e:
        result.status = "ERROR"
        result.error = str(e)
        result.duration = time.time() - start

    return result


def print_result(result: TestResult):
    """Print a test result."""
    icon = {"PASS": "✓", "FAIL": "✗", "ERROR": "!", "SKIP": "○"}.get(result.status, "?")
    color = {"PASS": "\033[32m", "FAIL": "\033[31m", "ERROR": "\033[33m", "SKIP": "\033[90m"}.get(result.status, "")
    reset = "\033[0m"

    print(f"  {color}{icon} {result.test_id}{reset} [{result.agent}] {result.duration:.1f}s — {result.status}")

    if result.status != "PASS":
        if result.error:
            print(f"    Error: {result.error}")
        for check, value in result.methodology_checks.items():
            if value in ("MISSING", "VIOLATION"):
                print(f"    {value}: {check}")


def generate_sample_files():
    """Generate sample files by executing code directly in the code-interpreter.
    Downloads produced files to tests/results/files/."""
    FILES_DIR.mkdir(parents=True, exist_ok=True)
    print(f"\n{'='*60}")
    print(f"Generating sample files via code-interpreter")
    print(f"Output: {FILES_DIR}/")
    print(f"{'='*60}\n")

    results = []
    for test_id, code in FILE_GENERATION_SCRIPTS.items():
        lang = "js" if "require(" in code else "py"
        print(f"  Generating {test_id} ({lang})...", end=" ")
        try:
            resp = exec_code(code.strip(), lang)
            stdout = resp.get("stdout", "")
            stderr = resp.get("stderr", "")
            files = resp.get("files", [])
            session_id = resp.get("session_id", "")

            if stderr and "Error" in stderr:
                print(f"ERROR: {stderr[:100]}")
                results.append((test_id, "ERROR", stderr[:100]))
                continue

            if not files:
                print(f"NO FILES (stdout: {stdout.strip()[:60]})")
                results.append((test_id, "NO_FILES", stdout.strip()[:60]))
                continue

            downloaded = []
            for f in files:
                file_id = f.get("id", "")
                file_name = f.get("name", "unknown")
                out_path = FILES_DIR / f"{test_id}_{file_name}"
                ok = download_file(session_id, file_id, out_path)
                if ok:
                    size = out_path.stat().st_size
                    downloaded.append(f"{file_name} ({size:,} bytes)")
                else:
                    downloaded.append(f"{file_name} (DOWNLOAD FAILED)")

            print(f"OK: {', '.join(downloaded)}")
            results.append((test_id, "OK", downloaded))

        except Exception as e:
            print(f"EXCEPTION: {e}")
            results.append((test_id, "EXCEPTION", str(e)))

    ok_count = sum(1 for _, s, _ in results if s == "OK")
    print(f"\n{'='*60}")
    print(f"Generated: {ok_count}/{len(results)} files")
    print(f"Files saved in: {FILES_DIR.absolute()}/")
    print(f"{'='*60}")

    if FILES_DIR.exists():
        for f in sorted(FILES_DIR.iterdir()):
            if f.is_file():
                print(f"  {f.name} ({f.stat().st_size:,} bytes)")


def main():
    parser = argparse.ArgumentParser(description="Run agent API tests")
    parser.add_argument("tests", nargs="*", help="Specific test IDs to run (e.g., D01b D01c P01b)")
    parser.add_argument("--agent", help="Run all tests for an agent (docx, pptx, xlsx, pdf, ffmpeg, dataviz)")
    parser.add_argument("--list", action="store_true", help="List all available tests")
    parser.add_argument("--key", help="Agent API key (overrides AGENT_API_KEY env var)")
    parser.add_argument("--jwt-token", help="LibreChat JWT token for file uploads")
    parser.add_argument("--no-file", action="store_true",
                        help="Skip tests that need input files (run self-contained only)")
    parser.add_argument("--only-file", action="store_true",
                        help="Run ONLY tests that need input files")
    parser.add_argument("--generate-files", action="store_true",
                        help="Generate sample files via code-interpreter (bypasses LLM)")
    args = parser.parse_args()

    global API_KEY, JWT_TOKEN
    if args.key:
        API_KEY = args.key
    if args.jwt_token:
        JWT_TOKEN = args.jwt_token
    if not API_KEY:
        # Try .agent-api-key file
        key_file = Path(__file__).parent.parent / ".agent-api-key"
        if key_file.exists():
            API_KEY = key_file.read_text().strip()
        if not API_KEY:
            # Try .env
            env_path = Path(__file__).parent.parent / ".env"
            if env_path.exists():
                for line in env_path.read_text().splitlines():
                    if line.startswith("AGENT_API_KEY="):
                        API_KEY = line.split("=", 1)[1].strip()
        if not API_KEY:
            print("ERROR: Set AGENT_API_KEY env var, pass --key, or put key in .agent-api-key")
            sys.exit(1)

    if args.list:
        total = len(TESTS)
        needs_file_count = sum(1 for t in TESTS if t.needs_file)
        self_contained = total - needs_file_count
        print(f"Available tests ({total} total: {self_contained} self-contained, "
              f"{needs_file_count} need files):\n")
        for t in TESTS:
            tags = []
            if t.test_id in FILE_GENERATION_SCRIPTS:
                tags.append("GEN")
            if t.needs_file:
                n_files = len(TEST_INPUT_FILES.get(t.test_id, []))
                tags.append(f"FILE:{n_files}")
            tag_str = f" [{','.join(tags)}]" if tags else ""
            print(f"  {t.test_id:6s} [{t.agent_name:8s}] {t.description}{tag_str}")
        print(f"\n[GEN] = has file generation script (--generate-files)")
        print(f"[FILE:N] = needs N input file(s) from docs/Test_files/")
        return

    if args.generate_files:
        generate_sample_files()
        return

    # Filter tests
    tests_to_run = TESTS
    if args.tests:
        tests_to_run = [t for t in TESTS if t.test_id in args.tests]
    elif args.agent:
        agent_map = {
            "docx": "DOCX", "pptx": "PPTX", "xlsx": "XLSX",
            "pdf": "PDF", "ffmpeg": "FFmpeg", "dataviz": "DataViz",
        }
        agent_name = agent_map.get(args.agent.lower(), args.agent)
        tests_to_run = [t for t in TESTS if t.agent_name == agent_name]

    # Apply file filter
    if args.no_file:
        tests_to_run = [t for t in tests_to_run if not t.needs_file]
    elif args.only_file:
        tests_to_run = [t for t in tests_to_run if t.needs_file]

    if not tests_to_run:
        print("No tests matched. Use --list to see available tests.")
        sys.exit(1)

    # Check file availability
    needs_file_tests = [t for t in tests_to_run if t.needs_file]
    if needs_file_tests:
        missing = 0
        for t in needs_file_tests:
            for rel_path in TEST_INPUT_FILES.get(t.test_id, []):
                if not (TEST_FILES_DIR / rel_path).exists():
                    missing += 1
        if missing > 0:
            print(f"WARNING: {missing} input file(s) missing from {TEST_FILES_DIR}/")
        if not JWT_TOKEN:
            print(f"INFO: No JWT token — file uploads disabled (methodology-only mode)")
            print(f"      Set LIBRECHAT_JWT or pass --jwt-token to enable file uploads")

    # Run tests
    print(f"\n{'='*60}")
    print(f"Running {len(tests_to_run)} agent tests")
    print(f"API: {API_BASE}")
    if JWT_TOKEN:
        print(f"File upload: enabled")
    print(f"{'='*60}\n")

    results = []
    for test_case in tests_to_run:
        file_tag = " [FILE]" if test_case.needs_file else ""
        print(f"  Running {test_case.test_id}{file_tag} ({test_case.description})...")
        result = run_test(test_case)
        results.append(result)
        print_result(result)
        print()

    # Summary
    passed = sum(1 for r in results if r.status == "PASS")
    failed = sum(1 for r in results if r.status == "FAIL")
    errors = sum(1 for r in results if r.status == "ERROR")
    skipped = sum(1 for r in results if r.status == "SKIP")
    total_time = sum(r.duration for r in results)

    print(f"{'='*60}")
    print(f"Results: {passed} passed, {failed} failed, {errors} errors / {len(results)} total")
    print(f"Total time: {total_time:.1f}s")
    print(f"{'='*60}")

    # Save results
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    report_path = RESULTS_DIR / f"report_{timestamp}.json"
    report = {
        "timestamp": timestamp,
        "total": len(results),
        "passed": passed,
        "failed": failed,
        "errors": errors,
        "skipped": skipped,
        "duration": total_time,
        "file_upload": "enabled" if JWT_TOKEN else "disabled",
        "results": [
            {
                "test_id": r.test_id,
                "agent": r.agent,
                "status": r.status,
                "duration": r.duration,
                "error": r.error,
                "files_uploaded": len(r.files_uploaded),
                "methodology_checks": r.methodology_checks,
                "response_preview": r.response_text[:200],
            }
            for r in results
        ],
    }
    report_path.write_text(json.dumps(report, indent=2, ensure_ascii=False))
    print(f"\nReport saved: {report_path}")

    sys.exit(1 if failed + errors > 0 else 0)


if __name__ == "__main__":
    main()
